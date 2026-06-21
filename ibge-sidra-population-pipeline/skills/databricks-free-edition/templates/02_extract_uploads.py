# Databricks notebook source
# EduLake BR-style: extracts zips and converts xlsx to CSV inside Unity Catalog Volumes.
# ⚠️ For files > 100MB, EXTRACT ON THE USER'S LOCAL MACHINE (see the
# "Robust extraction pattern" in the parent SKILL.md). This notebook only
# handles small files (Censo, IDEB, small xlsx) — for ENEM/CSVs > 100MB
# use `unzip` locally then upload the extracted CSVs to the Volume.
#
# Customize the SOURCES dict to match your project.
# Idempotent: skips files already extracted.
# Run as a Python notebook (dbutils is required for some ops).

import os
import io
import zipfile
import tempfile

# COMMAND ----------

# === CONFIGURE ===
# Map: (volume_path, [(source_zip/xlsx_path, file_type, extraction_strategy), ...])
# extraction_strategy: "first_csv", "all_csvs", "essential_csvs", "xlsx_to_csv"
SOURCES = {
    "enem": {
        "volume": "/Volumes/edulake/bronze/enem_raw/",
        "items": [
            # SKIP large zips — extract locally with `unzip` instead.
            # ("microdados_enem_2020.zip", "zip", "first_csv"),
        ],
    },
    "censo": {
        "volume": "/Volumes/edulake/bronze/censo_raw/",
        "items": [
            ("microdados_censo_escolar_2023.zip", "zip", "essential_csvs"),
        ],
    },
    "ideb": {
        "volume": "/Volumes/edulake/bronze/ideb_raw/",
        "items": [
            ("divulgacao_brasil_ideb_2023.zip", "zip", "first_csv_or_xlsx"),
        ],
    },
    "pib": {
        "volume": "/Volumes/edulake/bronze/pib_raw/",
        "items": [
            ("tabelas_completas_2022.xlsx", "xlsx", "xlsx_to_csv"),
        ],
    },
}

# COMMAND ----------


def file_exists(path):
    """Free Edition 2026+: only dbutils.fs.ls is available for existence checks."""
    try:
        dbutils.fs.ls(path)
        return True
    except Exception:
        return False


def write_bytes_via_tmp(path, data):
    """Write bytes to a Volume path. Free Edition 2026+ requires writing to
    /tmp first then dbutils.fs.cp('file:...', path). Direct dbutils.fs.open
    is NOT available (AttributeError: 'RemoteFsHandler' object has no attribute 'open')."""
    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        tmp.write(data)
        tmp_path = tmp.name
    try:
        dbutils.fs.cp(f"file:{tmp_path}", path)
    finally:
        os.unlink(tmp_path)


def read_bytes_spark(path):
    """Read bytes via Spark binaryFile. ⚠️ FAILS at 128MB gRPC limit on Free Edition.
    Only use for files < 100MB. For larger files, extract locally."""
    df = spark.read.format("binaryFile").load(path)
    return df.first().content


def extract_first_csv(zip_bytes, target_volume, zip_name):
    """Extract the first .csv from a zip into the Volume."""
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        csv_name = next(n for n in zf.namelist() if n.endswith(".csv") and not n.startswith("__MACOSX"))
        target = f"{target_volume}{os.path.basename(csv_name)}"
        if file_exists(target):
            return f"skipped (already extracted)"
        write_bytes_via_tmp(target, zf.open(csv_name).read())
        return f"extracted to {target}"


def extract_essential_csvs(zip_bytes, target_volume, keywords=["matriculas", "escolas", "docentes"]):
    """Extract CSVs whose name contains one of the keywords (saves storage)."""
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        results = []
        all_csvs = [n for n in zf.namelist() if n.endswith(".csv") and not n.startswith("__MACOSX")]
        essencial = [n for n in all_csvs if any(kw in n.lower() for kw in keywords)]
        if not essencial:
            essencial = all_csvs[:3]  # fallback: first 3
        for name in essencial:
            target = f"{target_volume}{os.path.basename(name)}"
            if file_exists(target):
                results.append(f"  skipped {os.path.basename(name)}")
                continue
            write_bytes_via_tmp(target, zf.open(name).read())
            results.append(f"  extracted {os.path.basename(name)}")
        return "\n".join(results)


def extract_first_csv_or_xlsx(zip_bytes, target_volume):
    """Pick the first .csv or .xlsx in the zip."""
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            if name.startswith("__MACOSX"):
                continue
            if name.endswith(".csv") or name.endswith(".xlsx"):
                target = f"{target_volume}{os.path.basename(name)}"
                if file_exists(target):
                    return f"skipped {os.path.basename(name)} (already extracted)"
                write_bytes_via_tmp(target, zf.open(name).read())
                return f"extracted {os.path.basename(name)}"


def convert_xlsx_to_csv(xlsx_bytes, target_volume, target_csv_name, sheet_keyword="per capita", header_row=0):
    """Convert IBGE-style XLSX (multi-sheet) to CSV via openpyxl. Picks sheet by keyword.
    Pass header_row=N if the file's header is at row N (not row 0)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    print(f"  Sheets in XLSX: {wb.sheetnames}")

    target_sheet = None
    for s in wb.sheetnames:
        if sheet_keyword.lower() in s.lower():
            target_sheet = s
            break
    if not target_sheet:
        target_sheet = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]

    print(f"  Converting sheet: {target_sheet} (using header at row {header_row})")
    target_path = f"{target_volume}{target_csv_name}"
    if file_exists(target_path):
        return f"skipped (CSV already exists at {target_path})"

    # Read with pandas (handles header offset better than openpyxl for non-trivial headers)
    import pandas as pd
    df = pd.read_excel(io.BytesIO(xlsx_bytes), sheet_name=target_sheet, header=header_row)
    print(f"  Sheet shape: {df.shape[0]} rows x {df.shape[1]} cols")
    print(f"  First 3 cols: {list(df.columns)[:3]}")

    # Encode as Latin-1 to match INEP/IBGE convention
    csv_data = df.to_csv(index=False, encoding="latin-1").encode("latin-1")
    write_bytes_via_tmp(target_path, csv_data)
    return f"converted to {target_path}"


# COMMAND ----------

# Run extraction for each source
for label, cfg in SOURCES.items():
    print(f"\n=== {label.upper()} ===")
    for filename, ftype, strategy in cfg["items"]:
        full_path = f"{cfg['volume']}{filename}"
        print(f"\nProcessing {filename} (strategy={strategy})...")
        try:
            file_bytes = read_bytes_spark(full_path)
            if ftype == "zip":
                if strategy == "first_csv":
                    result = extract_first_csv(file_bytes, cfg["volume"], filename)
                elif strategy == "essential_csvs":
                    result = extract_essential_csvs(file_bytes, cfg["volume"])
                elif strategy == "first_csv_or_xlsx":
                    result = extract_first_csv_or_xlsx(file_bytes, cfg["volume"])
                else:
                    result = f"unknown strategy: {strategy}"
            elif ftype == "xlsx":
                csv_name = filename.replace(".xlsx", ".csv")
                # Adjust header_row if the IBGE file has metadata at the top
                # PIB: try header=0 first; IDEB: usually header=2 or 3
                header_row = 2 if "ideb" in filename.lower() else 0
                result = convert_xlsx_to_csv(file_bytes, cfg["volume"], csv_name, header_row=header_row)
            else:
                result = f"unknown file type: {ftype}"
            print(f"  ✅ {result}")
        except Exception as e:
            print(f"  ❌ Error: {e}")

# COMMAND ----------

# Final validation: list each Volume's contents
print("\n=== Final state: contents of each Volume ===")
for label, cfg in SOURCES.items():
    print(f"\n{label} ({cfg['volume']}):")
    try:
        files = dbutils.fs.ls(cfg["volume"])
        for f in files:
            size_str = f"{f.size / 1024 / 1024:.1f} MB" if f.size > 1024 * 1024 else f"{f.size / 1024:.1f} KB"
            print(f"  📄 {f.name}  ({size_str})")
    except Exception as e:
        print(f"  ❌ Error: {e}")

print("\n✅ Extraction complete. Now run the 4 CREATE TABLE SQLs in SQL Editor.")
